from __future__ import annotations
import csv, io, logging, uuid, json, zipfile
from datetime import date, datetime, timezone

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import text

from backend.app.api.deps import get_db
from backend.app.api.schemas import ReportRequest, ReportResponse
from backend.app.analytics.queries import AnalyticsQueries

from backend.config.config import get_settings

router = APIRouter()
logger = logging.getLogger(__name__)

def _to_csv(rows: list, title: str = "") -> str:
    if not rows:
        return title + "\n(no data)\n"
    buf = io.StringIO()
    if title:
        buf.write(title + "\n")
    w = csv.DictWriter(buf, fieldnames=rows[0].keys())
    w.writeheader()
    w.writerows(rows)
    return buf.getvalue()

@router.post("/reports/generate", response_model=ReportResponse,
             summary="Згенерувати CSV/Excel звіт")
async def generate_report(body: dict,db: Session = Depends(get_db)):
    report_id = str(uuid.uuid4())[:8]
    sections = body.get("sections", [])
    start = body.get("start_date", "2025-01-01")
    end = body.get("end_date", "2025-12-31")
    fmt = body.get("format", "zip")
    return {
        "report_id": report_id,
        "status": "ready",
        "download_url": f"/api/v1/reports/{report_id}/download"
                        f"?start={start}&end={end}"
                        f"&sections={','.join(sections)}&fmt={fmt}",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }

@router.get("/reports/{report_id}/download", summary="Завантажити звіт CSV")
async def download_report(
    report_id: str,
    start: date = date(2025, 1, 1),
    end: date = date(2025, 12, 31),
    sections: str  = "",
    fmt: str  = "csv",
    db: Session = Depends(get_db),
):
    cfg = get_settings()
    q = AnalyticsQueries(db)
    sel = set(sections.split(",")) if sections else set()

    files: dict[str, str] = {}

    if not sel or "dashboard" in sel or "consumption" in sel:
        monthly = q.monthly_consumption(start.year)
        files["lr1_monthly_consumption.csv"] = _to_csv(
            monthly.to_dict(orient="records"),
            f"ЛР1: Місячне споживання {start.year}",
        )

    if not sel or "consumption" in sel:
        try:
            daily = q.daily_consumption(str(start), str(end))
            files["lr1_daily_consumption.csv"] = _to_csv(
                daily.to_dict(orient="records"),
                f"ЛР1: Добове споживання {start} – {end}",
            )
        except Exception as e:
            logger.warning("daily: %s", e)

    if not sel or "consumption" in sel:
        try:
            tariff = q.tariff_zone_analysis(str(start), str(end))
            files["lr1_tariff_zones.csv"] = _to_csv(
                tariff.to_dict(orient="records"),
                "ЛР1: Споживання по тарифних зонах",
            )
        except Exception as e:
            logger.warning("tariff: %s", e)

    if not sel or "forecast" in sel or "model_metrics" in sel:
        try:
            from backend.app.forecasting.forecast_service import ForecastService
            svc = ForecastService.from_saved(db)
            fc = svc.forecast_next_month()
            fc.index = fc.index.astype(str)
            files["lr2_forecast_hourly.csv"] = _to_csv(
                fc.reset_index().rename(columns={"index": "timestamp"}).to_dict(orient="records"),
                "ЛР2: Погодинний прогноз (наступний місяць)",
            )
        except Exception as e:
            logger.warning("forecast: %s", e)

    if not sel or "model_metrics" in sel:
        try:
            from backend.app.forecasting.forecast_service import ForecastService
            svc = ForecastService.from_saved(db)
            ts_pred = svc.predict_test_set()
            if not ts_pred.empty:
                ts_pred.index = ts_pred.index.astype(str)
                files["lr2_test_predictions.csv"] = _to_csv(
                    ts_pred.reset_index().rename(columns={"index": "timestamp"}).to_dict(orient="records"),
                    "ЛР2: Передбачення на тестовому наборі (Листопад-Грудень 2025)",
                )
        except Exception as e:
            logger.warning("test_pred: %s", e)

    if not sel or "ems_energy" in sel or "ems_economic" in sel or "energy_flow" in sel:
        try:
            rid_sql = text("SELECT run_id FROM simulation_runs ORDER BY started_at DESC LIMIT 1")
            row = db.execute(rid_sql).fetchone()
            run_id = row[0] if row else None
            if run_id:
                sim_sql = text("""
                    SELECT timestamp, solar_kwh, load_kwh, soc_pct, charge_kwh,
                           discharge_kwh, import_kwh, export_kwh,
                           tariff_zone, rate_uah_kwh, cost_uah
                    FROM simulation_results WHERE run_id = :rid ORDER BY timestamp
                """)
                rows = db.execute(sim_sql, {"rid": run_id}).fetchall()
                if rows:
                    files["lr3_ems_simulation.csv"] = _to_csv(
                        [dict(r._mapping) for r in rows],
                        f"ЛР3: EMS симуляція (run_id={run_id})",
                    )
        except Exception as e:
            logger.warning("ems_sim: %s", e)

    if not sel or "ems_economic" in sel:
        try:
            run_sql = text("""
                SELECT run_id, strategy, total_consumption_kwh, total_generation_kwh,
                       total_import_kwh, total_export_kwh, total_cost_uah,
                       baseline_cost_uah, savings_uah, simple_payback_years, npv_uah, irr_pct
                FROM simulation_runs WHERE status='completed' ORDER BY started_at DESC
            """)
            rows = db.execute(run_sql).fetchall()
            if rows:
                files["lr3_ems_runs.csv"] = _to_csv(
                    [dict(r._mapping) for r in rows],
                    "ЛР3: EMS прогони (усі завершені)",
                )
        except Exception as e:
            logger.warning("ems_runs: %s", e)

    lines = [
        "EMS — Energy Management System",
        f"Звіт згенеровано: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}",
        f"Об'єкт: {cfg.object.name}  |  Площа: {cfg.object.area_m2} м²",
        f"Режим: {cfg.object.operation_mode}  |  Тариф: день {cfg.tariff.zones[0].rate_uah_kwh if cfg.tariff.zones else '6.9'} / ніч 5.6",
        "",
    ]
    if "lr1_monthly_consumption.csv" in files:
        try:
            monthly = q.monthly_consumption(start.year)
            lines += [
                "ЛР1: Аналітика споживання",
                f"Річне споживання: {monthly['total_kwh'].sum():,.0f} кВт·год",
                f"Річна вартість: {monthly['total_cost_uah'].sum():,.0f} грн",
                f"Питоме: {monthly['total_kwh'].sum()/cfg.object.area_m2:.1f} кВт·год/м²",
                "",
            ]
        except Exception: pass
    if "lr2_forecast_hourly.csv" in files:
        try:
            svc = ForecastService.from_saved(db)
            s = svc.forecast_summary()
            lines += [
                "ЛР2: Прогнозування (Gradient Boosting)",
                f"Прогноз місяця: {s['monthly_kwh']:,.0f} кВт·год",
                f"Вартість: {s['monthly_cost_uah']:,.0f} грн",
                f"Економія СЕС: {s['solar_saving_uah']:,.0f} грн",
                f"Питоме: {s['specific_kwh_m2']} кВт·год/м²",
                "",
            ]
        except Exception: pass
    try:
        run_sql = text("SELECT total_cost_uah, baseline_cost_uah, savings_uah, simple_payback_years, npv_uah, irr_pct FROM simulation_runs WHERE status='completed' ORDER BY started_at DESC LIMIT 1")
        r = db.execute(run_sql).fetchone()
        if r:
            lines += [
                "ЛР3: EMS Engine",
                f"Вартість з EMS: {r.total_cost_uah:,.0f} грн",
                f"Вартість без EMS: {r.baseline_cost_uah:,.0f} грн",
                f"Економія: {r.savings_uah:,.0f} грн",
                f"Simple Payback: {r.simple_payback_years:.1f} р.",
                f"NPV (r=12%): {r.npv_uah:,.0f} грн",
                f"IRR: {r.irr_pct:.1f}%",
                "",
            ]
    except Exception: pass
    lines += [
        f"Файли у архіві: {', '.join(files.keys())}",
    ]
    files["summary.txt"] = "\n".join(lines)

    if fmt in ("zip", "csv"):
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for name, content in files.items():
                zf.writestr(name, content.encode("utf-8"))
        zip_buf.seek(0)
        filename = f"ems_report_{report_id}_{start}_{end}.zip"
        return StreamingResponse(
            iter([zip_buf.read()]),
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    if fmt == "excel":
        try:
            import openpyxl
            from io import BytesIO
            wb = openpyxl.Workbook()
            first = True
            for fname, content in files.items():
                if fname.endswith(".txt"):
                    continue
                ws = wb.active if first else wb.create_sheet()
                ws.title = fname.replace(".csv","")[:31]
                first = False
                for i, line in enumerate(content.splitlines(), 1):
                    ws.cell(i, 1, line)
            buf = BytesIO(); wb.save(buf); buf.seek(0)
            return StreamingResponse(
                iter([buf.read()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename=ems_report_{report_id}.xlsx"},
            )
        except ImportError:
            raise HTTPException(400, "pip install openpyxl")

    raise HTTPException(400, f"Формат {fmt} не підтримується")

@router.get("/reports/{report_id}/preview")
async def preview_report(
    report_id: str,
    start: date = date(2025, 1, 1),
    end: date = date(2025, 12, 31),
    sections: str  = "",
    db: Session = Depends(get_db),
):
    cfg = get_settings()
    q = AnalyticsQueries(db)
    lines = [
        "EMS — Energy Management System  |  ПОПЕРЕДНІЙ ПЕРЕГЛЯД ЗВІТУ",
        f"Період: {start} – {end}",
        f"Об'єкт: {cfg.object.name}  |  Площа: {cfg.object.area_m2} м²",
        f"Обрані розділи: {sections or 'усі'}",
    ]
    try:
        monthly = q.monthly_consumption(start.year)
        lines += [
            "ЛР1: Споживання",
            f"Річне споживання: {monthly['total_kwh'].sum():,.0f} кВт·год",
            f"Річна вартість: {monthly['total_cost_uah'].sum():,.0f} грн",
            f"Питоме: {monthly['total_kwh'].sum()/cfg.object.area_m2:.1f} кВт·год/м²", "",
        ]
    except Exception: pass
    try:
        from backend.app.forecasting.forecast_service import ForecastService
        svc = ForecastService.from_saved(db)
        s   = svc.forecast_summary()
        lines += [
            "ЛР2: Прогноз (Gradient Boosting)",
            f"Прогноз місяця: {s['monthly_kwh']:,.0f} кВт·год",
            f"Вартість: {s['monthly_cost_uah']:,.0f} грн",
            f"Економія СЕС: {s['solar_saving_uah']:,.0f} грн", "",
        ]
    except Exception: pass
    try:
        r = db.execute(text("SELECT total_cost_uah, baseline_cost_uah, savings_uah, simple_payback_years, npv_uah, irr_pct FROM simulation_runs WHERE status='completed' ORDER BY started_at DESC LIMIT 1")).fetchone()
        if r:
            lines += [
                "ЛР3: EMS",
                f"Вартість з EMS: {r.total_cost_uah:,.0f} грн",
                f"Вартість без EMS: {r.baseline_cost_uah:,.0f} грн",
                f"Економія: {r.savings_uah:,.0f} грн",
                f"Payback: {r.simple_payback_years:.1f} р.   NPV: {r.npv_uah:,.0f} грн   IRR: {r.irr_pct:.1f}%", "",
            ]
    except Exception: pass
    lines += ["Архів міститиме: CSV-таблиці + summary.txt"]
    from fastapi.responses import PlainTextResponse
    return PlainTextResponse("\n".join(lines))
